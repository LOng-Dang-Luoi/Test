from PyQt6 import QtCore,QtGui,QtWidgets,uic
from PyQt6.QtGui import QKeySequence
from PyQt6.QtWidgets import *
from PyQt6.uic import loadUi
import sys
import openpyxl
from random import randint,choice
import time
from playsound import playsound
#Thiết lập trình phát âm lượng
import pygame
pygame.mixer.init()
pygame.mixer.music.set_volume(0.5)


import pyttsx3
engine = pyttsx3.init()

#Biến lưu tên trang tính    
name_sheet = ''
#Lưu số bắt đầu và kết thúc của người dùng trong giao diện 2
num_start = 1     
num_end = 50
#List chức các stt câu hỏi
Question_list = []  

# -------------> Screen 1 
class Main(QMainWindow):
    def __init__(self):
        super(Main,self).__init__()
        uic.loadUi('file_gui/login.ui',self)
        self.setWindowTitle("Start Learing")
        #Lấy name tất cả trang tính trong file và ghi lên giao diện 1
        workbook = openpyxl.load_workbook("data.xlsx")
        sheet_names = workbook.sheetnames
        for i in sheet_names:
            self.from_data.addItem(i)
        #Chuyển hướng tới giao diện 2
        self.start.setShortcut(QKeySequence("ctrl+t"))
        self.start.clicked.connect(self.continu)
    def continu(self):
        pygame.mixer.music.load("img/button.wav")
        pygame.mixer.music.play()
        global name_sheet
        wiget.setCurrentIndex(1)
        name_sheet = self.from_data.currentText()
        S_data.show(name_sheet)

#----------------> Screen 2
class Show_Data(QMainWindow):
    def __init__(self):
        super(Show_Data,self).__init__()
        uic.loadUi('file_gui/show_data.ui',self)
        #Thiết lập độ rộng cho từng cột
        self.save.setShortcut(QKeySequence("ctrl+s"))
        self.default_.setShortcut(QKeySequence("ctrl+d"))
        self.remove_row.setShortcut(QKeySequence("ctrl+x"))
        self.learing.setShortcut(QKeySequence("ctrl+t"))

        self.tableWidget.setColumnWidth(0, 160)
        self.tableWidget.setColumnWidth(1, 110)
        self.tableWidget.setColumnWidth(2, 70)
        self.tableWidget.setColumnWidth(3, 150)
        self.save.clicked.connect(self.saved)
        self.default_.clicked.connect(self.defau)
        self.learing.clicked.connect(self.Learing)
        self.home.clicked.connect(self.back_home)
        self.remove_row.clicked.connect(self.remove)
        self.add_row.clicked.connect(self.add)
    def saved(self):
        num_start = int(self.start_point.toPlainText())
        num_end = int(self.end_point.toPlainText())
        number_of_rows = self.tableWidget.rowCount()
        c=0
        while self.tableWidget.rowCount()!=num_end:
            self.tableWidget.removeRow(num_end)
            c+=1
            if c>10000:
                break
        c=0
        while c!=num_start-1:
            self.tableWidget.removeRow(0)
            c+=1
        self.start_point.setText("1")
        self.end_point.setText(str(num_end-num_start+1))
    def remove(self):
        row = self.tableWidget.currentRow()
        self.tableWidget.removeRow(row)
        number_of_rows = self.tableWidget.rowCount()
        self.start_point.setText("1")
        self.end_point.setText(str(number_of_rows))
    def add(self):
        row = self.tableWidget.currentRow()
        self.tableWidget.insertRow(row+1)
        number_of_rows = self.tableWidget.rowCount()
        self.start_point.setText("1")
        self.end_point.setText(str(number_of_rows))
    def clear(self):
        while self.tableWidget.rowCount()!=1:
            self.tableWidget.removeRow(0)
    def back_home(self):
        self.clear()
        wiget.setCurrentIndex(0)

    #Ghi dữ liệu ra bảng
    def show(self,name_sheet: str):
        #Get
        self.name_sheet = name_sheet
        self.wb = openpyxl.load_workbook("data.xlsx")
        #Mở trang tính
        self.sheet = self.wb[self.name_sheet]
        #Lấy số dòng
        self.num_row = self.sheet.max_row

        self.start_point.setText("1")
        self.end_point.setText(str(self.num_row))

        number_of_rows = self.tableWidget.rowCount()
        self.tableWidget.setHorizontalHeaderLabels(["Từ Vựng", "Phiên Âm" ,"Loại Từ" ,"Dịch Nghĩa" ,"Đánh Giá"])
        self.tableWidget.setItem(0, 0, QTableWidgetItem("--"))
        for  i in range(0,self.num_row):
            current_row = self.tableWidget.currentRow()
            self.tableWidget.insertRow(current_row)
        self.tableWidget.removeRow(self.num_row)
        for i in range(0,self.num_row):
            tu_vung = self.sheet.cell(row=i+1,column=1).value
            phien_am = self.sheet.cell(row=i+1,column=2).value
            loai_tu = self.sheet.cell(row=i+1,column=3).value
            dich_nghia = self.sheet.cell(row=i+1,column=4).value
            self.tableWidget.setItem(i, 0, QTableWidgetItem(tu_vung))
            self.tableWidget.setItem(i, 1, QTableWidgetItem(phien_am))
            self.tableWidget.setItem(i, 2, QTableWidgetItem(loai_tu))
            self.tableWidget.setItem(i, 3, QTableWidgetItem(dich_nghia))
    def defau(self):
        self.clear()
        self.show(name_sheet)
    #Chuyển hướng tới giao diện học(Learning)
    def Learing(self):
        self.saved()
        pygame.mixer.music.load("img/button.wav")
        pygame.mixer.music.play()
        global num_start,num_end, Dict_data
        num_start = int(self.start_point.toPlainText())
        num_end = int(self.end_point.toPlainText())
        
        #List chức các json mỗi từ !!!!!      -Importand
        Dict_data = []
        for i in range(0,num_end):
            dt = {}
            tu_vung = self.tableWidget.item(i, 0).text()
            loai_tu = self.tableWidget.item(i, 1).text()
            phien_am = self.tableWidget.item(i, 2).text()
            nghia = self.tableWidget.item(i, 3).text()
            dt['tu_vung'] = tu_vung
            dt['loai_tu'] = loai_tu
            dt['phien_am'] = phien_am
            dt['nghia'] = nghia
            Dict_data.append(dt)
        wiget.setCurrentIndex(2)
        Learn.run()

def load(Dict_data):
    pass

#---------------> Screen 3
class Learning(QMainWindow):
    def __init__(self):
        
        super(Learning,self).__init__()
        uic.loadUi('file_gui/learning.ui',self)
        self.num=0
        self.okey.clicked.connect(self.check)
        self.kq=0
        self.text.setFocus()
        self.setWindowTitle("Learning")
        self.okey.setShortcut(QKeySequence("Return"))
        self.voice.clicked.connect(self.voi)
    def voi(self):
        playsound(f'data/{name_sheet}/{self.tu_vung.text().lower()}.mp3')
    #Hàm show câu hỏi ra windowns
    def show(self,tu_vung,phien_am,loai_tu,nghia):
        self.text.setStyleSheet("border-radius: 8px;font: 75 12pt \"MS Shell Dlg 2\";background-color:  rgb(255, 255, 255);")
        self.tu_vung.setText(tu_vung.title())
        self.loai_tu.setText(loai_tu.title())
        self.phien_am.setText(phien_am)
        self.nghia.setText(nghia)
        self.text.setFocus()
        
    def run(self):
        self.text.setText('')
        question = Dict_data[self.num]
        tu_vung = question['tu_vung']
        loai_tu = question['loai_tu']
        phien_am = question['phien_am']
        nghia = question['nghia']
        self.show(tu_vung,phien_am,loai_tu,nghia)
        self.name = tu_vung

    def cotinu(self):
        self.num += 1
        self.kq = 0
    def win_lose(self,kq):
        if kq == True:
            self.text.setStyleSheet("border-radius: 8px;font: 75 12pt \"MS Shell Dlg 2\";background-color: rgb(60, 182, 89);")
            pygame.mixer.music.load("img/correct.wav")
            pygame.mixer.music.play()
        else:
            self.text.setStyleSheet("border-radius: 8px;font: 75 12pt \"MS Shell Dlg 2\";background-color:  rgb(212, 0, 0);")
            pygame.mixer.music.load("img/notcorrect.wav")
            pygame.mixer.music.play()
        # time.sleep(1)
    def check(self):
        playsound(f'data/{name_sheet}/{self.tu_vung.text().lower()}.mp3')
        kq = None
        print(self.tu_vung.text().lower(),self.text.text().lower())
        if self.tu_vung.text().lower() == self.text.text().lower():
            self.kq +=1
            kq = True
            print('Yess')
        else:
            kq = False
            print('Noo')

        if self.kq == 5:
            self.cotinu()
        self.win_lose(kq)
        self.run()

#Thiết lập chính
app=QApplication(sys.argv)
wiget=QtWidgets.QStackedWidget()

#Khởi tạo các màn hình
starts=Main()
S_data=Show_Data()
Learn = Learning()

#Thêm từng màn hình và bắt đầu với screen 1
wiget.addWidget(starts)
wiget.addWidget(S_data)
wiget.addWidget(Learn)
wiget.setCurrentIndex(0)
wiget.setFixedHeight(480)   #xét chiều cao
wiget.setFixedWidth(640)   #xét chiều rộng
wiget.show()
app.exec()